from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import date, timedelta
import json, csv, io
from decimal import Decimal
from .models import Sale, SaleItem, SalePayment, Customer, Expense, DailySummary, CashDrawerSession, CreditAccount, CreditTransaction, CreditTransactionItem
from inventory.models import Product, StockMovement
from accounts.decorators import module_required, write_required, admin_required


# Valid methods for a single payment line (excludes the composite 'split')
PAYMENT_LINE_METHODS = {m for m, _ in Sale.PAYMENT_LINE_METHOD}


def _parse_payments(raw_payments, total, fallback_method='cash'):
    """Normalise an incoming list of payment dicts into [(method, Decimal)].

    Accepts a list like [{'method': 'cash', 'amount': '500'}, ...]. When the
    list is empty/None (e.g. legacy offline sales that only sent a single
    payment_method), a single line for the full total is synthesised.

    Returns (payments, error). `payments` is a list of (method, Decimal) and
    `error` is a string when validation fails (or None).
    """
    total = Decimal(str(total or 0))
    payments = []
    for p in (raw_payments or []):
        method = (p.get('method') or '').strip()
        if method not in PAYMENT_LINE_METHODS:
            return None, f'Invalid payment method "{method}".'
        try:
            amount = Decimal(str(p.get('amount', 0) or 0))
        except (ValueError, ArithmeticError):
            return None, 'Invalid payment amount.'
        if amount <= 0:
            continue
        payments.append((method, amount))

    if not payments:
        # No explicit payment lines — treat the whole total as one method.
        method = fallback_method if fallback_method in PAYMENT_LINE_METHODS else 'cash'
        return [(method, total)], None

    paid = sum(a for _, a in payments)
    if abs(paid - total) > Decimal('0.01'):
        return None, f'Payments (₦{paid}) must equal the sale total (₦{total}).'
    return payments, None


def _record_payments(sale, payments):
    """Create SalePayment rows and set the sale's payment_method label."""
    for method, amount in payments:
        SalePayment.objects.create(sale=sale, method=method, amount=amount)
    sale.payment_method = payments[0][0] if len(payments) == 1 else 'split'
    sale.save(update_fields=['payment_method'])


def _payment_breakdown(sales, methods):
    """Sum each sale's payment_list into a {method: float} breakdown."""
    breakdown = {k: 0 for k in methods}
    for s in sales:
        for method, amount in s.payment_list:
            if method in breakdown:
                breakdown[method] += float(amount)
    return breakdown


@login_required
@module_required('sales')
def sale_list(request):
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', '')
    payment = request.GET.get('payment', '')
    q = request.GET.get('q', '')

    sales = Sale.objects.select_related('customer', 'served_by').order_by('-sale_date')

    if date_from:
        sales = sales.filter(sale_date__date__gte=date_from)
    if date_to:
        sales = sales.filter(sale_date__date__lte=date_to)
    if status:
        sales = sales.filter(status=status)
    if payment:
        sales = sales.filter(payment_method=payment)
    if q:
        sales = sales.filter(Q(invoice_number__icontains=q) | Q(customer__name__icontains=q))

    # Paginate before grouping
    paginator = Paginator(sales, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_sales = list(page_obj)

    # Group by date
    today = date.today()
    yesterday = today - timedelta(days=1)
    groups = {}
    for sale in page_sales:
        d = sale.sale_date.date()
        groups.setdefault(d, []).append(sale)
    grouped_sales = sorted(groups.items(), reverse=True)

    total_revenue = sum(s.total_amount for s in sales if s.status == 'completed')
    return render(request, 'sales/sale_list.html', {
        'grouped_sales': grouped_sales,
        'page_obj': page_obj,
        'total_revenue': total_revenue,
        'today': today,
        'yesterday': yesterday,
        'date_from': date_from, 'date_to': date_to,
        'status': status, 'payment': payment, 'q': q,
    })


@login_required
@module_required('pos')
def pos_view(request):
    from inventory.models import Category
    products = Product.objects.filter(is_active=True, quantity_in_stock__gt=0).select_related('category')
    categories = Category.objects.filter(
        products__is_active=True, products__quantity_in_stock__gt=0
    ).distinct().order_by('name')
    return render(request, 'sales/pos.html', {
        'products': products,
        'categories': categories,
    })


@login_required
@module_required('pos')
def new_sale(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        discount = Decimal(str(request.POST.get('discount_amount', 0) or 0))
        notes = request.POST.get('notes', '')
        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')
        prices = request.POST.getlist('unit_price')

        # Payment lines — supports split payments via parallel POST lists.
        pay_methods = request.POST.getlist('payment_method')
        pay_amounts = request.POST.getlist('payment_amount')
        raw_payments = [{'method': m, 'amount': a}
                        for m, a in zip(pay_methods, pay_amounts)]

        customer = None
        if customer_id:
            try:
                customer = Customer.objects.get(pk=customer_id)
            except Customer.DoesNotExist:
                pass

        # Validate stock before creating the sale
        line_items = []
        for pid, qty, price in zip(product_ids, quantities, prices):
            if not pid or not qty:
                continue
            try:
                product = Product.objects.get(pk=pid, is_active=True)
                qty_int = int(qty)
                if qty_int < 1:
                    messages.error(request, f'Invalid quantity for product ID {pid}.')
                    return redirect('new_sale')
                if product.quantity_in_stock < qty_int:
                    messages.error(request, f'Insufficient stock for "{product.name}". Available: {product.quantity_in_stock}.')
                    return redirect('new_sale')
                line_items.append((product, qty_int, Decimal(str(price))))
            except (Product.DoesNotExist, ValueError):
                messages.error(request, f'Invalid product or quantity.')
                return redirect('new_sale')

        if not line_items:
            messages.error(request, 'No valid items in sale.')
            return redirect('new_sale')

        sale_total = sum(qty * price for _, qty, price in line_items) - discount
        payments, pay_error = _parse_payments(raw_payments, sale_total)
        if pay_error:
            messages.error(request, pay_error)
            return redirect('new_sale')

        sale = Sale.objects.create(
            customer=customer,
            served_by=request.user,
            discount_amount=discount,
            notes=notes,
            status='completed',
            payment_status=True,
        )
        _record_payments(sale, payments)

        for product, qty_int, price in line_items:
            SaleItem.objects.create(
                sale=sale, product=product,
                quantity=qty_int,
                unit_price=price,
                unit_cost=product.cost_price,
            )
            prev = product.quantity_in_stock
            product.quantity_in_stock = max(0, product.quantity_in_stock - qty_int)
            product.save()
            StockMovement.objects.create(
                product=product, movement_type='out',
                quantity=qty_int, previous_stock=prev,
                new_stock=product.quantity_in_stock,
                reference=sale.invoice_number,
                created_by=request.user,
            )

        messages.success(request, f'Sale {sale.invoice_number} recorded successfully.')
        return redirect('sale_detail', pk=sale.pk)

    customers = Customer.objects.all().order_by('name')
    products = Product.objects.filter(is_active=True, quantity_in_stock__gt=0)
    return render(request, 'sales/new_sale.html', {
        'customers': customers, 'products': products,
    })


@login_required
@module_required('sales')
def sale_detail(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'sales/sale_detail.html', {'sale': sale})


@login_required
@module_required('sales')
@write_required
def sale_edit(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    if request.method == 'POST':
        sale.status = request.POST.get('status', sale.status)
        sale.payment_method = request.POST.get('payment_method', sale.payment_method)
        sale.discount_amount = Decimal(str(request.POST.get('discount_amount', sale.discount_amount) or 0))
        sale.notes = request.POST.get('notes', sale.notes)
        sale.save()
        messages.success(request, f'Sale {sale.invoice_number} updated.')
        return redirect('sale_detail', pk=sale.pk)
    return render(request, 'sales/sale_edit.html', {'sale': sale})


@login_required
@module_required('sales')
@write_required
def sale_delete(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    if request.method == 'POST':
        # Restore stock for each item
        for item in sale.items.all():
            product = item.product
            prev = product.quantity_in_stock
            product.quantity_in_stock += item.quantity
            product.save()
            StockMovement.objects.create(
                product=product, movement_type='adjustment',
                quantity=item.quantity, previous_stock=prev,
                new_stock=product.quantity_in_stock,
                reference=f'VOID-{sale.invoice_number}',
                created_by=request.user,
            )
        sale.delete()
        messages.success(request, 'Sale deleted and stock restored.')
        return redirect('sale_list')
    return render(request, 'sales/sale_confirm_delete.html', {'sale': sale})


@login_required
@module_required('sales')
def sale_invoice(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'sales/invoice.html', {'sale': sale})


@login_required
@module_required('customers')
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    sales = customer.sales.order_by('-sale_date')
    total_spent = sum(s.total_amount for s in sales if s.status == 'completed')
    return render(request, 'sales/customer_detail.html', {
        'customer': customer, 'sales': sales, 'total_spent': total_spent,
    })


@login_required
@module_required('customers')
@write_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.name = request.POST.get('name')
        customer.email = request.POST.get('email', '')
        customer.phone = request.POST.get('phone', '')
        customer.address = request.POST.get('address', '')
        customer.save()
        messages.success(request, 'Customer updated.')
        return redirect('customer_detail', pk=customer.pk)
    return render(request, 'sales/customer_form.html', {'customer': customer})


@login_required
@module_required('customers')
@write_required
def customer_delete(request, pk):
    if request.method == 'POST':
        customer = get_object_or_404(Customer, pk=pk)
        customer.delete()
        messages.success(request, 'Customer deleted.')
    return redirect('customer_list')


@login_required
@module_required('customers')
def customer_list(request):
    customers = Customer.objects.all().order_by('name')
    q = request.GET.get('q', '')
    if q:
        customers = customers.filter(Q(name__icontains=q) | Q(phone__icontains=q) | Q(email__icontains=q))
    return render(request, 'sales/customer_list.html', {'customers': customers, 'q': q})


@login_required
@module_required('customers')
@write_required
def customer_add(request):
    if request.method == 'POST':
        Customer.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            address=request.POST.get('address', ''),
        )
        messages.success(request, 'Customer added.')
        return redirect('customer_list')
    return render(request, 'sales/customer_form.html')


@login_required
@module_required('expenses')
def expense_list(request):
    expenses = Expense.objects.order_by('-expense_date')
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    if month:
        year, m = month.split('-')
        expenses = expenses.filter(expense_date__year=year, expense_date__month=m)
    total = expenses.aggregate(t=Sum('amount'))['t'] or 0
    return render(request, 'sales/expense_list.html', {
        'expenses': expenses, 'total': total, 'month': month,
    })


@login_required
@module_required('expenses')
@write_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.title = request.POST.get('title')
        expense.category = request.POST.get('category')
        expense.amount = request.POST.get('amount')
        expense.description = request.POST.get('description', '')
        expense.expense_date = request.POST.get('expense_date') or date.today()
        expense.save()
        messages.success(request, 'Expense updated.')
        return redirect('expense_list')
    return render(request, 'sales/expense_form.html', {'expense': expense})


@login_required
@module_required('expenses')
@write_required
def expense_delete(request, pk):
    if request.method == 'POST':
        expense = get_object_or_404(Expense, pk=pk)
        expense.delete()
        messages.success(request, 'Expense deleted.')
    return redirect('expense_list')


@login_required
@module_required('expenses')
@write_required
def expense_add(request):
    if request.method == 'POST':
        Expense.objects.create(
            title=request.POST.get('title'),
            category=request.POST.get('category'),
            amount=request.POST.get('amount'),
            description=request.POST.get('description', ''),
            expense_date=request.POST.get('expense_date', date.today()),
            recorded_by=request.user,
        )
        messages.success(request, 'Expense recorded.')
        return redirect('expense_list')
    return render(request, 'sales/expense_form.html')


@login_required
@module_required('reports')
def reports_view(request):
    today = date.today()
    month_start = today.replace(day=1)
    period = request.GET.get('period', 'month')

    if period == 'today':
        start_date = today
    elif period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = month_start
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = month_start

    sales = Sale.objects.filter(sale_date__date__gte=start_date, status='completed').prefetch_related('payments')
    total_revenue = sum(s.total_amount for s in sales)
    total_profit = sum(s.total_profit for s in sales)
    total_transactions = sales.count()
    expenses = Expense.objects.filter(expense_date__gte=start_date)
    total_expenses = expenses.aggregate(t=Sum('amount'))['t'] or 0
    net_profit = total_profit - total_expenses
    net_revenue = total_revenue - total_expenses

    # Sales by payment method (split-payment aware) — drop empty buckets
    payment_breakdown = {
        method: amount
        for method, amount in _payment_breakdown(sales, PAYMENT_LINE_METHODS).items()
        if amount > 0
    }

    # Top selling products
    from sales.models import SaleItem
    from django.db.models import ExpressionWrapper, DecimalField, F as Fref
    top_products = SaleItem.objects.filter(
        sale__sale_date__date__gte=start_date,
        sale__status='completed'
    ).values('product__name').annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(Fref('quantity') * Fref('unit_price'), output_field=DecimalField()))
    ).order_by('-total_qty')[:10]

    # Daily chart
    chart_data = []
    days = (today - start_date).days + 1
    for i in range(min(days, 30)):
        d = today - timedelta(days=(min(days, 30) - 1 - i))
        day_sales = Sale.objects.filter(sale_date__date=d, status='completed')
        day_rev = sum(s.total_amount for s in day_sales)
        chart_data.append({'date': d.strftime('%d %b'), 'revenue': float(day_rev)})

    context = {
        'total_revenue': total_revenue,
        'total_profit': total_profit,
        'total_transactions': total_transactions,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'net_revenue': net_revenue,
        'payment_breakdown': payment_breakdown,
        'top_products': top_products,
        'chart_data': chart_data,
        'period': period,
        'start_date': start_date,
    }
    return render(request, 'sales/reports.html', context)


@login_required
@module_required('pos')
def product_search_api(request):
    q = request.GET.get('q', '')
    products = Product.objects.filter(
        Q(name__icontains=q) | Q(sku__icontains=q),
        is_active=True, quantity_in_stock__gt=0
    )[:10]
    data = [{
        'id': p.id,
        'name': p.name,
        'sku': p.sku,
        'price': float(p.effective_price),
        'stock': p.quantity_in_stock,
        'image': p.image.url if p.image else '',
    } for p in products]
    return JsonResponse({'products': data})


@login_required
@module_required('pos')
def process_sale_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            if not items:
                return JsonResponse({'success': False, 'error': 'No items in sale.'}, status=400)
            discount = Decimal(str(data.get('discount', 0) or 0))
            customer_name = data.get('customer_name', '').strip()[:200]

            # Validate stock before touching anything
            products_to_update = []
            sale_total = Decimal('0')
            for item in items:
                try:
                    product = Product.objects.get(pk=item['id'])
                except Product.DoesNotExist:
                    return JsonResponse({'success': False, 'error': f'Product ID {item["id"]} not found.'}, status=400)
                qty = int(item['quantity'])
                if qty < 1:
                    return JsonResponse({'success': False, 'error': f'Invalid quantity for "{product.name}".'}, status=400)
                if product.quantity_in_stock < qty:
                    return JsonResponse({'success': False, 'error': f'Insufficient stock for "{product.name}". Available: {product.quantity_in_stock}.'}, status=400)
                products_to_update.append((product, qty, item['price']))
                sale_total += Decimal(str(item['price'])) * qty
            sale_total -= discount

            # Payment lines — `payments` is preferred; fall back to the legacy
            # single `payment_method` field (used by older offline-cached sales).
            payments, pay_error = _parse_payments(
                data.get('payments'), sale_total,
                fallback_method=data.get('payment_method', 'cash'),
            )
            if pay_error:
                return JsonResponse({'success': False, 'error': pay_error}, status=400)

            customer = None
            if customer_name:
                customer, _ = Customer.objects.get_or_create(
                    name=customer_name,
                    defaults={'is_walk_in': True}
                )

            sale = Sale.objects.create(
                customer=customer,
                served_by=request.user,
                discount_amount=discount,
                status='completed',
                payment_status=True,
            )
            _record_payments(sale, payments)

            for product, qty, price in products_to_update:
                SaleItem.objects.create(
                    sale=sale, product=product,
                    quantity=qty,
                    unit_price=Decimal(str(price)),
                    unit_cost=product.cost_price,
                )
                prev = product.quantity_in_stock
                product.quantity_in_stock = max(0, product.quantity_in_stock - qty)
                product.save()
                StockMovement.objects.create(
                    product=product, movement_type='out',
                    quantity=qty, previous_stock=prev,
                    new_stock=product.quantity_in_stock,
                    reference=sale.invoice_number,
                    created_by=request.user,
                )

            return JsonResponse({
                'success': True,
                'invoice_number': sale.invoice_number,
                'total': float(sale.total_amount),
                'sale_id': sale.pk,
            })
        except (ValueError, KeyError):
            return JsonResponse({'success': False, 'error': 'Invalid request data.'}, status=400)
        except Exception:
            return JsonResponse({'success': False, 'error': 'An error occurred processing the sale.'}, status=500)
    return JsonResponse({'success': False, 'error': 'Method not allowed.'}, status=405)


def _parse_upload(file):
    name = file.name.lower()
    if name.endswith('.csv'):
        decoded = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        return list(reader)
    elif name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))})
        return result
    return []


@login_required
@module_required('imports')
def import_customers(request):
    HEADERS = ['Name', 'Email', 'Phone', 'Address']

    if request.GET.get('template') == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='166534')
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="customers_template.xlsx"'
        return resp

    if request.GET.get('template') == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="customers_template.csv"'
        csv.writer(resp).writerow(HEADERS)
        return resp

    results = []
    if request.method == 'POST' and request.FILES.get('file'):
        rows = _parse_upload(request.FILES['file'])
        created = errors = 0
        for i, row in enumerate(rows, start=2):
            try:
                name = row.get('Name', '').strip()
                if not name:
                    continue
                Customer.objects.create(
                    name=name,
                    email=row.get('Email', ''),
                    phone=row.get('Phone', ''),
                    address=row.get('Address', ''),
                )
                results.append({'row': i, 'name': name, 'status': 'created'})
                created += 1
            except Exception as e:
                results.append({'row': i, 'name': row.get('Name', f'Row {i}'), 'status': 'error', 'error': str(e)})
                errors += 1
        messages.success(request, f'Import complete: {created} customer(s) added, {errors} error(s).')

    return render(request, 'sales/import_customers.html', {
        'entity': 'Customers', 'headers': HEADERS,
        'results': results,
    })


@login_required
@module_required('pos')
def sale_receipt(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'sales/pos_receipt.html', {'sale': sale})


@login_required
@module_required('reports')
def daily_report(request):
    today = date.today()

    # Default to current month
    month = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        year, m = map(int, month.split('-'))
        period_start = date(year, m, 1)
        if m == 12:
            period_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            period_end = date(year, m + 1, 1) - timedelta(days=1)
    except (ValueError, TypeError):
        period_start = today.replace(day=1)
        period_end = today
        month = today.strftime('%Y-%m')

    # Build one row per active day
    days_range = (period_end - period_start).days + 1
    all_days = [period_start + timedelta(days=i) for i in range(days_range)]

    # Fetch all completed sales and expenses in one shot
    sales_qs = Sale.objects.filter(
        sale_date__date__gte=period_start,
        sale_date__date__lte=period_end,
        status='completed',
    ).prefetch_related('payments')

    expenses_qs = Expense.objects.filter(
        expense_date__gte=period_start,
        expense_date__lte=period_end,
    )

    # Index by date for fast lookup
    sales_by_date = {}
    for s in sales_qs:
        d = s.sale_date.date()
        sales_by_date.setdefault(d, []).append(s)

    expenses_by_date = {}
    for e in expenses_qs:
        expenses_by_date.setdefault(e.expense_date, []).append(e)

    payment_labels = {
        'cash': 'Cash',
        'card': 'Card',
        'transfer': 'Transfer',
        'pos': 'POS',
        'online': 'Online',
    }

    daily_rows = []
    period_totals = {k: 0 for k in ['transactions', 'revenue', 'cash', 'card', 'transfer', 'pos', 'online', 'expenses', 'net']}

    for d in all_days:
        day_sales = sales_by_date.get(d, [])
        day_expenses = expenses_by_date.get(d, [])
        if not day_sales and not day_expenses:
            continue

        breakdown = {k: 0 for k in payment_labels}
        revenue = 0
        for s in day_sales:
            revenue += float(s.total_amount)
            for method, amount in s.payment_list:
                if method in breakdown:
                    breakdown[method] += float(amount)

        exp_total = sum(float(e.amount) for e in day_expenses)
        net = revenue - exp_total

        daily_rows.append({
            'date': d,
            'transactions': len(day_sales),
            'revenue': revenue,
            'breakdown': breakdown,
            'expenses': exp_total,
            'net': net,
        })

        period_totals['transactions'] += len(day_sales)
        period_totals['revenue'] += revenue
        for k in payment_labels:
            period_totals[k] += breakdown[k]
        period_totals['expenses'] += exp_total
        period_totals['net'] += net

    # Chart: daily revenue for the period in chronological order (oldest → newest)
    chart_data = [{'date': r['date'].strftime('%d %b'), 'revenue': r['revenue']} for r in daily_rows]

    # Table lists the most recent day first so the current day is at the top.
    daily_rows = list(reversed(daily_rows))

    return render(request, 'sales/daily_report.html', {
        'daily_rows': daily_rows,
        'period_totals': period_totals,
        'chart_data': json.dumps(chart_data),
        'month': month,
        'period_start': period_start,
        'period_end': period_end,
        'payment_labels': payment_labels,
    })


@login_required
@module_required('reports')
def daily_report_detail(request, report_date):
    try:
        from datetime import datetime
        target = datetime.strptime(report_date, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('daily_report')

    sales = Sale.objects.filter(
        sale_date__date=target,
        status='completed',
    ).select_related('customer', 'served_by').prefetch_related('items__product', 'payments').order_by('sale_date')

    expenses = Expense.objects.filter(expense_date=target).order_by('created_at')

    payment_labels = {
        'cash': 'Cash',
        'card': 'Card',
        'transfer': 'Transfer',
        'pos': 'POS',
        'online': 'Online',
    }

    breakdown = {k: 0 for k in payment_labels}
    revenue = 0
    for s in sales:
        revenue += float(s.total_amount)
        for method, amount in s.payment_list:
            if method in breakdown:
                breakdown[method] += float(amount)

    exp_total = sum(float(e.amount) for e in expenses)
    net = revenue - exp_total

    summary = {
        'revenue': revenue,
        'transactions': sales.count(),
        'breakdown': breakdown,
        'expenses': exp_total,
        'net': net,
    }

    prev_date = target - timedelta(days=1)
    next_date = target + timedelta(days=1)
    today = date.today()

    return render(request, 'sales/daily_report_detail.html', {
        'target': target,
        'sales': sales,
        'expenses': expenses,
        'summary': summary,
        'payment_labels': payment_labels,
        'prev_date': prev_date,
        'next_date': next_date if next_date <= today else None,
    })


# ── Cash drawer (per cashier, per day) ──────────────────────────────────────

def _user_can_view_all_cash(user):
    if user.is_superuser:
        return True
    try:
        return user.staff_profile.has_module_access('reports')
    except Exception:
        return False


def _suggested_opening_float(user, day):
    """Carry forward the closing float from this cashier's most recent
    closed session before `day`."""
    prev = CashDrawerSession.objects.filter(
        cashier=user, date__lt=day, is_closed=True
    ).order_by('-date').first()
    return prev.closing_float if prev else Decimal('0')


@login_required
@module_required('pos')
def cash_drawer(request):
    today = timezone.localdate()
    session, created = CashDrawerSession.objects.get_or_create(
        cashier=request.user, date=today,
        defaults={'opening_float': _suggested_opening_float(request.user, today)},
    )

    my_sessions = CashDrawerSession.objects.filter(
        cashier=request.user
    ).exclude(pk=session.pk).order_by('-date')[:30]

    can_view_all = _user_can_view_all_cash(request.user)
    all_sessions = None
    if can_view_all:
        all_sessions = CashDrawerSession.objects.select_related('cashier').filter(
            date=today
        ).exclude(pk=session.pk).order_by('cashier__username')

    return render(request, 'sales/cash_drawer.html', {
        'session': session,
        'my_sessions': my_sessions,
        'all_sessions': all_sessions,
        'can_view_all': can_view_all,
        'today': today,
    })


@login_required
@module_required('pos')
def cash_drawer_open(request):
    if request.method == 'POST':
        today = timezone.localdate()
        session, _ = CashDrawerSession.objects.get_or_create(
            cashier=request.user, date=today,
            defaults={'opening_float': _suggested_opening_float(request.user, today)},
        )
        if session.is_closed:
            messages.error(request, 'Today\'s drawer is already closed and cannot be changed.')
            return redirect('cash_drawer')
        try:
            session.opening_float = Decimal(str(request.POST.get('opening_float', 0) or 0))
        except (ValueError, ArithmeticError):
            messages.error(request, 'Invalid opening float amount.')
            return redirect('cash_drawer')
        session.save(update_fields=['opening_float', 'updated_at'])
        messages.success(request, 'Opening cash balance saved.')
    return redirect('cash_drawer')


@login_required
@module_required('reports')
def cash_drawer_overview(request):
    """Admin view: all cashiers' cash sessions for a given date."""
    today = timezone.localdate()
    date_str = request.GET.get('date', today.isoformat())
    try:
        from datetime import datetime as _dt
        target_date = _dt.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        target_date = today

    sessions = (
        CashDrawerSession.objects
        .filter(date=target_date)
        .select_related('cashier')
        .order_by('cashier__first_name', 'cashier__username')
    )

    totals = {
        'opening': sum(s.opening_float for s in sessions),
        'cash_sales': sum(s.cash_sales for s in sessions),
        'expected': sum(s.expected_cash for s in sessions),
        'transferred': sum(s.cash_transferred_to_bank for s in sessions),
        'closing': sum(s.closing_float for s in sessions),
    }

    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)

    return render(request, 'sales/cash_drawer_overview.html', {
        'sessions': sessions,
        'target_date': target_date,
        'date_str': date_str,
        'today': today,
        'totals': totals,
        'prev_date': prev_date,
        'next_date': next_date if next_date <= today else None,
    })


@login_required
@module_required('pos')
def cash_drawer_close(request):
    if request.method == 'POST':
        today = timezone.localdate()
        session = get_object_or_404(CashDrawerSession, cashier=request.user, date=today)
        if session.is_closed:
            messages.error(request, 'Today\'s drawer is already closed.')
            return redirect('cash_drawer')
        try:
            transferred = Decimal(str(request.POST.get('cash_transferred_to_bank', 0) or 0))
            counted = Decimal(str(request.POST.get('counted_cash', 0) or 0))
        except (ValueError, ArithmeticError):
            messages.error(request, 'Invalid amount entered.')
            return redirect('cash_drawer')

        session.cash_transferred_to_bank = transferred
        session.counted_cash = counted
        # Cash carried forward to the next day = expected on hand − transferred.
        session.closing_float = session.expected_cash - transferred
        session.notes = request.POST.get('notes', '')[:1000]
        session.is_closed = True
        session.closed_at = timezone.now()
        session.save()
        messages.success(
            request,
            f'Day closed. ₦{session.closing_float} carried forward to the next day.'
        )
    return redirect('cash_drawer')


# ── Credit accounts ──────────────────────────────────────────────────────────

@login_required
@module_required('customers')
def credit_list(request):
    accounts = (
        CreditAccount.objects
        .select_related('customer')
        .prefetch_related('transactions')
        .order_by('customer__name')
    )
    return render(request, 'sales/credit_list.html', {'accounts': accounts})


@login_required
@module_required('customers')
def credit_detail(request, pk):
    account = get_object_or_404(
        CreditAccount.objects.select_related('customer').prefetch_related('transactions'),
        pk=pk,
    )
    transactions = account.transactions.prefetch_related('items__product').order_by('date', 'created_at')

    running = Decimal('0')
    ledger = []
    for txn in transactions:
        if txn.transaction_type == 'issue':
            running += txn.amount
        else:
            running -= txn.amount
        ledger.append({'txn': txn, 'balance': running})
    ledger.reverse()

    return render(request, 'sales/credit_detail.html', {
        'account': account,
        'ledger': ledger,
        'today': date.today(),
    })


@login_required
@module_required('customers')
@write_required
def credit_account_create(request, customer_pk):
    customer = get_object_or_404(Customer, pk=customer_pk)
    if hasattr(customer, 'credit_account'):
        messages.info(request, f'{customer.name} already has a credit account.')
        return redirect('credit_detail', pk=customer.credit_account.pk)
    if request.method == 'POST':
        account = CreditAccount.objects.create(
            customer=customer,
            notes=request.POST.get('notes', ''),
            created_by=request.user,
        )
        messages.success(request, f'Credit account created for {customer.name}.')
        return redirect('credit_detail', pk=account.pk)
    return render(request, 'sales/credit_account_form.html', {'customer': customer})


@login_required
@module_required('customers')
@write_required
def credit_issue(request, account_pk):
    account = get_object_or_404(CreditAccount, pk=account_pk)
    if not account.is_active:
        messages.error(request, 'This credit account is inactive.')
        return redirect('credit_detail', pk=account.pk)

    if request.method == 'POST':
        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')
        prices = request.POST.getlist('unit_price')
        date_str = request.POST.get('date', '')
        notes = request.POST.get('notes', '')

        line_items = []
        for pid, qty, price in zip(product_ids, quantities, prices):
            if not pid or not qty:
                continue
            try:
                product = Product.objects.get(pk=pid, is_active=True)
                qty_int = int(qty)
                if qty_int < 1:
                    messages.error(request, f'Invalid quantity for {product.name}.')
                    return redirect('credit_issue', account_pk=account_pk)
                if product.quantity_in_stock < qty_int:
                    messages.error(request, f'Insufficient stock for "{product.name}". Available: {product.quantity_in_stock}.')
                    return redirect('credit_issue', account_pk=account_pk)
                line_items.append((product, qty_int, Decimal(str(price))))
            except (Product.DoesNotExist, ValueError):
                messages.error(request, 'Invalid product or quantity.')
                return redirect('credit_issue', account_pk=account_pk)

        if not line_items:
            messages.error(request, 'No valid items to issue.')
            return redirect('credit_issue', account_pk=account_pk)

        total = sum(qty * price for _, qty, price in line_items)

        try:
            issue_date = date.fromisoformat(date_str) if date_str else date.today()
        except ValueError:
            issue_date = date.today()

        txn = CreditTransaction.objects.create(
            account=account,
            transaction_type='issue',
            date=issue_date,
            amount=total,
            notes=notes,
            recorded_by=request.user,
        )

        for product, qty_int, price in line_items:
            CreditTransactionItem.objects.create(
                transaction=txn, product=product,
                quantity=qty_int, unit_price=price,
            )
            prev = product.quantity_in_stock
            product.quantity_in_stock = max(0, product.quantity_in_stock - qty_int)
            product.save()
            StockMovement.objects.create(
                product=product, movement_type='out',
                quantity=qty_int, previous_stock=prev,
                new_stock=product.quantity_in_stock,
                reference=txn.reference,
                notes=f'Credit issuance to {account.customer.name}',
                created_by=request.user,
            )

        messages.success(request, f'Goods issued. Reference: {txn.reference}')
        return redirect('credit_detail', pk=account.pk)

    products = Product.objects.filter(is_active=True).order_by('name')
    return render(request, 'sales/credit_issue.html', {
        'account': account,
        'products': products,
        'today': date.today().isoformat(),
    })


@login_required
@module_required('customers')
@write_required
def credit_payment(request, account_pk):
    account = get_object_or_404(CreditAccount, pk=account_pk)
    if request.method == 'POST':
        try:
            amount = Decimal(str(request.POST.get('amount', 0) or 0))
        except (ValueError, ArithmeticError):
            messages.error(request, 'Invalid amount.')
            return redirect('credit_detail', pk=account.pk)

        if amount <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            return redirect('credit_detail', pk=account.pk)

        date_str = request.POST.get('date', '')
        try:
            pay_date = date.fromisoformat(date_str) if date_str else date.today()
        except ValueError:
            pay_date = date.today()

        txn = CreditTransaction.objects.create(
            account=account,
            transaction_type='payment',
            date=pay_date,
            amount=amount,
            notes=request.POST.get('notes', ''),
            recorded_by=request.user,
        )
        messages.success(request, f'Payment of ₦{amount:,.2f} recorded. Ref: {txn.reference}')
    return redirect('credit_detail', pk=account.pk)


@login_required
@module_required('customers')
@write_required
def credit_return(request, account_pk):
    account = get_object_or_404(CreditAccount, pk=account_pk)
    if not account.is_active:
        messages.error(request, 'This credit account is inactive.')
        return redirect('credit_detail', pk=account.pk)

    if request.method == 'POST':
        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')
        prices = request.POST.getlist('unit_price')
        date_str = request.POST.get('date', '')
        notes = request.POST.get('notes', '')

        line_items = []
        for pid, qty, price in zip(product_ids, quantities, prices):
            if not pid or not qty:
                continue
            try:
                product = Product.objects.get(pk=pid, is_active=True)
                qty_int = int(qty)
                if qty_int < 1:
                    messages.error(request, f'Invalid quantity for {product.name}.')
                    return redirect('credit_return', account_pk=account_pk)
                line_items.append((product, qty_int, Decimal(str(price))))
            except (Product.DoesNotExist, ValueError):
                messages.error(request, 'Invalid product or quantity.')
                return redirect('credit_return', account_pk=account_pk)

        if not line_items:
            messages.error(request, 'No valid items to return.')
            return redirect('credit_return', account_pk=account_pk)

        total = sum(qty * price for _, qty, price in line_items)

        try:
            ret_date = date.fromisoformat(date_str) if date_str else date.today()
        except ValueError:
            ret_date = date.today()

        txn = CreditTransaction.objects.create(
            account=account,
            transaction_type='return',
            date=ret_date,
            amount=total,
            notes=notes,
            recorded_by=request.user,
        )

        for product, qty_int, price in line_items:
            CreditTransactionItem.objects.create(
                transaction=txn, product=product,
                quantity=qty_int, unit_price=price,
            )
            prev = product.quantity_in_stock
            product.quantity_in_stock += qty_int
            product.save()
            StockMovement.objects.create(
                product=product, movement_type='return',
                quantity=qty_int, previous_stock=prev,
                new_stock=product.quantity_in_stock,
                reference=txn.reference,
                notes=f'Credit return from {account.customer.name}',
                created_by=request.user,
            )

        messages.success(request, f'Return recorded. Reference: {txn.reference}')
        return redirect('credit_detail', pk=account.pk)

    products = Product.objects.filter(is_active=True).order_by('name')
    return render(request, 'sales/credit_return.html', {
        'account': account,
        'products': products,
        'today': date.today().isoformat(),
    })
