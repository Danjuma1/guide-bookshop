from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, F
from django.utils.text import slugify
from django.http import HttpResponse
from .models import Product, Category, Supplier, StockMovement, PurchaseOrder, PurchaseOrderItem
from accounts.decorators import module_required, write_required
import uuid
import csv
import io


@login_required
@module_required('inventory')
def product_list(request):
    products = Product.objects.select_related('category', 'supplier').filter(is_active=True)
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    product_type = request.GET.get('type', '')

    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q) | Q(author__icontains=q))
    if category_id:
        products = products.filter(category_id=category_id)
    if product_type:
        products = products.filter(product_type=product_type)

    categories = Category.objects.all()
    return render(request, 'inventory/product_list.html', {
        'products': products,
        'categories': categories,
        'q': q,
        'selected_category': category_id,
        'selected_type': product_type,
    })


@login_required
@module_required('inventory')
@write_required
def product_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        slug = slugify(name) + '-' + str(uuid.uuid4())[:4]
        product = Product(
            name=name, slug=slug,
            sku=request.POST.get('sku') or f"SKU-{str(uuid.uuid4())[:8].upper()}",
            product_type=request.POST.get('product_type', 'book'),
            category_id=request.POST.get('category') or None,
            supplier_id=request.POST.get('supplier') or None,
            description=request.POST.get('description', ''),
            author=request.POST.get('author', ''),
            isbn=request.POST.get('isbn', ''),
            publisher=request.POST.get('publisher', ''),
            edition=request.POST.get('edition', ''),
            cost_price=request.POST.get('cost_price', 0),
            selling_price=request.POST.get('selling_price', 0),
            discount_price=request.POST.get('discount_price') or None,
            quantity_in_stock=request.POST.get('quantity_in_stock', 0),
            reorder_level=request.POST.get('reorder_level', 5),
            is_featured='is_featured' in request.POST,
            is_available_online='is_available_online' in request.POST,
        )
        if request.FILES.get('image'):
            product.image = request.FILES['image']
        product.save()

        # Record initial stock movement
        if int(request.POST.get('quantity_in_stock', 0)) > 0:
            StockMovement.objects.create(
                product=product,
                movement_type='in',
                quantity=product.quantity_in_stock,
                previous_stock=0,
                new_stock=product.quantity_in_stock,
                reference='Initial Stock',
                created_by=request.user
            )

        messages.success(request, f'Product "{name}" added successfully.')
        return redirect('product_list')

    categories = Category.objects.all()
    suppliers = Supplier.objects.filter(is_active=True)
    return render(request, 'inventory/product_form.html', {
        'categories': categories, 'suppliers': suppliers, 'action': 'Add'
    })


@login_required
@module_required('inventory')
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    movements = product.movements.order_by('-created_at')[:10]
    return render(request, 'inventory/product_detail.html', {
        'product': product, 'movements': movements
    })


@login_required
@module_required('inventory')
@write_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.name = request.POST.get('name')
        product.product_type = request.POST.get('product_type', 'book')
        product.category_id = request.POST.get('category') or None
        product.supplier_id = request.POST.get('supplier') or None
        product.description = request.POST.get('description', '')
        product.author = request.POST.get('author', '')
        product.isbn = request.POST.get('isbn', '')
        product.publisher = request.POST.get('publisher', '')
        product.edition = request.POST.get('edition', '')
        product.cost_price = request.POST.get('cost_price', 0)
        product.selling_price = request.POST.get('selling_price', 0)
        product.discount_price = request.POST.get('discount_price') or None
        product.reorder_level = request.POST.get('reorder_level', 5)
        product.is_featured = 'is_featured' in request.POST
        product.is_available_online = 'is_available_online' in request.POST
        product.is_active = 'is_active' in request.POST
        if request.FILES.get('image'):
            product.image = request.FILES['image']
        product.save()
        messages.success(request, 'Product updated.')
        return redirect('product_detail', pk=pk)

    categories = Category.objects.all()
    suppliers = Supplier.objects.filter(is_active=True)
    return render(request, 'inventory/product_form.html', {
        'product': product, 'categories': categories,
        'suppliers': suppliers, 'action': 'Edit'
    })


@login_required
@module_required('inventory')
@write_required
def adjust_stock(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        movement_type = request.POST.get('movement_type')
        quantity = int(request.POST.get('quantity', 0))
        notes = request.POST.get('notes', '')
        previous_stock = product.quantity_in_stock

        if movement_type == 'in':
            product.quantity_in_stock += quantity
        elif movement_type in ['out', 'damage']:
            product.quantity_in_stock = max(0, product.quantity_in_stock - quantity)
        elif movement_type == 'adjustment':
            product.quantity_in_stock = quantity

        product.save()
        StockMovement.objects.create(
            product=product, movement_type=movement_type,
            quantity=quantity, previous_stock=previous_stock,
            new_stock=product.quantity_in_stock,
            notes=notes, created_by=request.user
        )
        messages.success(request, f'Stock adjusted for {product.name}.')
        return redirect('product_detail', pk=pk)

    return render(request, 'inventory/adjust_stock.html', {'product': product})


@login_required
@module_required('inventory')
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'inventory/category_list.html', {'categories': categories})


@login_required
@module_required('inventory')
@write_required
def category_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        slug = slugify(name) + '-' + str(uuid.uuid4())[:4]
        Category.objects.create(
            name=name, slug=slug,
            description=request.POST.get('description', ''),
            parent_id=request.POST.get('parent') or None
        )
        messages.success(request, f'Category "{name}" added.')
        return redirect('category_list')
    categories = Category.objects.filter(parent=None)
    return render(request, 'inventory/category_form.html', {'categories': categories})


@login_required
@module_required('inventory')
@write_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description', '')
        parent_id = request.POST.get('parent') or None
        category.parent_id = parent_id if parent_id and int(parent_id) != pk else None
        category.save()
        messages.success(request, f'Category "{category.name}" updated.')
        return redirect('category_list')
    categories = Category.objects.filter(parent=None).exclude(pk=pk)
    return render(request, 'inventory/category_form.html', {'category': category, 'categories': categories, 'action': 'Edit'})


@login_required
@module_required('inventory')
@write_required
def product_delete(request, pk):
    if request.method == 'POST':
        product = get_object_or_404(Product, pk=pk)
        name = product.name
        product.delete()
        messages.success(request, f'Product "{name}" deleted.')
    return redirect('product_list')


@login_required
@module_required('inventory')
def supplier_list(request):
    suppliers = Supplier.objects.filter(is_active=True)
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})


@login_required
@module_required('inventory')
@write_required
def supplier_add(request):
    if request.method == 'POST':
        Supplier.objects.create(
            name=request.POST.get('name'),
            contact_person=request.POST.get('contact_person', ''),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            address=request.POST.get('address', ''),
        )
        messages.success(request, 'Supplier added.')
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_form.html')


@login_required
@module_required('inventory')
@write_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.name = request.POST.get('name')
        supplier.contact_person = request.POST.get('contact_person', '')
        supplier.email = request.POST.get('email', '')
        supplier.phone = request.POST.get('phone', '')
        supplier.address = request.POST.get('address', '')
        supplier.is_active = 'is_active' in request.POST
        supplier.save()
        messages.success(request, 'Supplier updated.')
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_form.html', {'supplier': supplier, 'action': 'Edit'})


@login_required
@module_required('inventory')
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    products = supplier.product_set.filter(is_active=True).order_by('name')
    return render(request, 'inventory/supplier_detail.html', {
        'supplier': supplier, 'products': products,
    })


@login_required
@module_required('inventory')
@write_required
def supplier_delete(request, pk):
    if request.method == 'POST':
        supplier = get_object_or_404(Supplier, pk=pk)
        name = supplier.name
        supplier.delete()
        messages.success(request, f'Supplier "{name}" deleted.')
    return redirect('supplier_list')


@login_required
@module_required('inventory')
def stock_movement_list(request):
    movements = StockMovement.objects.select_related('product', 'created_by').order_by('-created_at')[:100]
    return render(request, 'inventory/stock_movements.html', {'movements': movements})


@login_required
@module_required('inventory')
def po_list(request):
    orders = PurchaseOrder.objects.select_related('supplier').order_by('-created_at')
    return render(request, 'inventory/po_list.html', {'orders': orders})


@login_required
@module_required('inventory')
@write_required
def po_add(request):
    if request.method == 'POST':
        import uuid as _uuid
        po_number = f"PO-{_uuid.uuid4().hex[:8].upper()}"
        po = PurchaseOrder.objects.create(
            po_number=po_number,
            supplier_id=request.POST.get('supplier'),
            expected_date=request.POST.get('expected_date') or None,
            notes=request.POST.get('notes', ''),
            created_by=request.user
        )
        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')
        costs = request.POST.getlist('unit_cost')
        for pid, qty, cost in zip(product_ids, quantities, costs):
            if pid and qty and cost:
                PurchaseOrderItem.objects.create(
                    purchase_order=po, product_id=pid,
                    quantity_ordered=int(qty), unit_cost=float(cost)
                )
        messages.success(request, f'Purchase Order {po_number} created.')
        return redirect('po_list')

    suppliers = Supplier.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True)
    return render(request, 'inventory/po_form.html', {'suppliers': suppliers, 'products': products})


@login_required
@module_required('inventory')
def po_detail(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST' and request.POST.get('action') == 'receive':
        for item in po.items.all():
            qty_received = int(request.POST.get(f'received_{item.pk}', 0))
            if qty_received > 0:
                prev = item.product.quantity_in_stock
                item.product.quantity_in_stock += qty_received
                item.product.cost_price = item.unit_cost
                item.product.save()
                item.quantity_received += qty_received
                item.save()
                StockMovement.objects.create(
                    product=item.product, movement_type='in',
                    quantity=qty_received, previous_stock=prev,
                    new_stock=item.product.quantity_in_stock,
                    reference=po.po_number, created_by=request.user
                )
        all_received = all(item.quantity_received >= item.quantity_ordered for item in po.items.all())
        po.status = 'received' if all_received else 'partial'
        po.save()
        messages.success(request, 'Stock received and updated.')
        return redirect('po_detail', pk=pk)
    return render(request, 'inventory/po_detail.html', {'po': po})


@login_required
@module_required('inventory')
def low_stock_report(request):
    low_stock = Product.objects.filter(
        quantity_in_stock__lte=F('reorder_level'), is_active=True
    ).order_by('quantity_in_stock')
    return render(request, 'inventory/low_stock.html', {'products': low_stock})


def _parse_upload(file):
    """Read an uploaded CSV or XLSX file and return list of row dicts."""
    name = file.name.lower()
    if name.endswith('.csv'):
        decoded = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        return list(reader)
    elif name.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))})
        return result
    return []


def _csv_template_response(filename, headers):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    return response


def _xlsx_template_response(filename, headers):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='166534')
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@module_required('imports')
def import_products(request):
    HEADERS = ['Name', 'SKU', 'Product Type', 'Category', 'Author', 'ISBN',
               'Publisher', 'Edition', 'Cost Price', 'Selling Price', 'Discount Price',
               'Quantity', 'Reorder Level', 'Description']

    if request.GET.get('template') == 'xlsx':
        return _xlsx_template_response('products_template.xlsx', HEADERS)
    if request.GET.get('template') == 'csv':
        return _csv_template_response('products_template.csv', HEADERS)

    results = []
    if request.method == 'POST' and request.FILES.get('file'):
        rows = _parse_upload(request.FILES['file'])
        created = errors = 0
        for i, row in enumerate(rows, start=2):
            try:
                name = row.get('Name', '').strip()
                if not name:
                    continue
                cat = None
                cat_name = row.get('Category', '').strip()
                if cat_name:
                    cat, _ = Category.objects.get_or_create(
                        name=cat_name,
                        defaults={'slug': slugify(cat_name) + '-' + str(uuid.uuid4())[:4]}
                    )
                slug = slugify(name) + '-' + str(uuid.uuid4())[:4]
                sku = row.get('SKU', '').strip() or f"SKU-{str(uuid.uuid4())[:8].upper()}"
                if Product.objects.filter(sku=sku).exists():
                    sku = f"SKU-{str(uuid.uuid4())[:8].upper()}"
                qty = int(float(row.get('Quantity', 0) or 0))
                product = Product.objects.create(
                    name=name, slug=slug, sku=sku,
                    product_type=row.get('Product Type', 'book').lower().replace(' ', '_') or 'book',
                    category=cat,
                    author=row.get('Author', ''),
                    isbn=row.get('ISBN', ''),
                    publisher=row.get('Publisher', ''),
                    edition=row.get('Edition', ''),
                    cost_price=float(row.get('Cost Price', 0) or 0),
                    selling_price=float(row.get('Selling Price', 0) or 0),
                    discount_price=float(row.get('Discount Price') or 0) or None,
                    quantity_in_stock=qty,
                    reorder_level=int(float(row.get('Reorder Level', 5) or 5)),
                    description=row.get('Description', ''),
                )
                if qty > 0:
                    StockMovement.objects.create(
                        product=product, movement_type='in',
                        quantity=qty, previous_stock=0, new_stock=qty,
                        reference='Bulk Import', created_by=request.user
                    )
                results.append({'row': i, 'name': name, 'status': 'created'})
                created += 1
            except Exception as e:
                results.append({'row': i, 'name': row.get('Name', f'Row {i}'), 'status': 'error', 'error': str(e)})
                errors += 1
        messages.success(request, f'Import complete: {created} product(s) created, {errors} error(s).')

    return render(request, 'inventory/import.html', {
        'entity': 'Products', 'headers': HEADERS,
        'template_csv': '?template=csv', 'template_xlsx': '?template=xlsx',
        'results': results,
        'upload_url': request.path,
    })


@login_required
@module_required('imports')
def import_suppliers(request):
    HEADERS = ['Name', 'Contact Person', 'Email', 'Phone', 'Address']

    if request.GET.get('template') == 'xlsx':
        return _xlsx_template_response('suppliers_template.xlsx', HEADERS)
    if request.GET.get('template') == 'csv':
        return _csv_template_response('suppliers_template.csv', HEADERS)

    results = []
    if request.method == 'POST' and request.FILES.get('file'):
        rows = _parse_upload(request.FILES['file'])
        created = errors = 0
        for i, row in enumerate(rows, start=2):
            try:
                name = row.get('Name', '').strip()
                if not name:
                    continue
                Supplier.objects.create(
                    name=name,
                    contact_person=row.get('Contact Person', ''),
                    email=row.get('Email', ''),
                    phone=row.get('Phone', ''),
                    address=row.get('Address', ''),
                )
                results.append({'row': i, 'name': name, 'status': 'created'})
                created += 1
            except Exception as e:
                results.append({'row': i, 'name': row.get('Name', f'Row {i}'), 'status': 'error', 'error': str(e)})
                errors += 1
        messages.success(request, f'Import complete: {created} supplier(s) created, {errors} error(s).')

    return render(request, 'inventory/import.html', {
        'entity': 'Suppliers', 'headers': HEADERS,
        'template_csv': '?template=csv', 'template_xlsx': '?template=xlsx',
        'results': results,
        'upload_url': request.path,
    })
